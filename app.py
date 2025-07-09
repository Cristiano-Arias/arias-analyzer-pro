import os
import tempfile
import shutil
import json
import logging
import re
from datetime import datetime
from typing import Dict, List, Any, Optional

# Flask e dependências web
from flask import Flask, request, render_template_string, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename

# Azure Document Intelligence
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.ai.documentintelligence.models import AnalyzeDocumentRequest
from azure.core.credentials import AzureKeyCredential

# Processamento de documentos (SEM PANDAS)
import PyPDF2
from openpyxl import load_workbook

# Geração de relatórios
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Configuração de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)

# Configurações
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf', 'xlsx', 'xls'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max

# Configurações Azure
AZURE_ENDPOINT = "https://proposal-analyzer-di.cognitiveservices.azure.com/"
AZURE_KEY = "SUA_CHAVE_AZURE_AQUI"

# Criar diretório de upload se não existir
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

class AzureDocumentIntelligenceExtractor:
    """Extrator usando Azure Document Intelligence para PDFs complexos"""
    
    def __init__(self, endpoint: str, key: str):
        self.endpoint = endpoint
        self.key = key
        try:
            self.client = DocumentIntelligenceClient(
                endpoint=endpoint,
                credential=AzureKeyCredential(key)
            )
            self.azure_available = True
            logger.info("Azure Document Intelligence inicializado com sucesso")
        except Exception as e:
            logger.error(f"Erro ao inicializar Azure: {str(e)}")
            self.azure_available = False
    
    def extract_from_pdf(self, pdf_path: str) -> Dict[str, Any]:
        """Extrai dados estruturados de PDF usando Azure Document Intelligence"""
        if not self.azure_available:
            return self._fallback_extraction(pdf_path)
        
        try:
            logger.info(f"Iniciando extração Azure para: {pdf_path}")
            
            with open(pdf_path, "rb") as f:
                pdf_content = f.read()
            
            # Analisar documento usando Layout model
            poller = self.client.begin_analyze_document(
                "prebuilt-layout",
                analyze_request=pdf_content,
                content_type="application/pdf"
            )
            
            result = poller.result()
            extracted_data = self._parse_azure_result(result)
            
            logger.info(f"Extração Azure concluída: {extracted_data['confidence_score']:.1f}% confiança")
            return extracted_data
            
        except Exception as e:
            logger.error(f"Erro na extração Azure: {str(e)}")
            return self._fallback_extraction(pdf_path)
    
    def _parse_azure_result(self, result) -> Dict[str, Any]:
        """Processa resultado do Azure Document Intelligence"""
        extracted_data = {
            'empresa': '',
            'cnpj': '',
            'metodologia': '',
            'metodologia_detalhada': '',
            'prazo_dias': 0,
            'cronograma_fases': [],
            'equipe_total': 0,
            'equipe_detalhada': {},
            'equipamentos': [],
            'materiais': [],
            'tecnologias': [],
            'arquitetura_sistema': '',
            'requisitos_tecnicos': [],
            'riscos_tecnicos': [],
            'preco_total': 0.0,
            'bdi_percentual': 0.0,
            'condicoes_pagamento': '',
            'garantia': '',
            'composicao_custos': {
                'mao_obra': 0.0,
                'materiais': 0.0,
                'equipamentos': 0.0
            },
            'raw_text': '',
            'tabelas': [],
            'confidence_score': 0.0
        }
        
        try:
            # Extrair texto completo
            if result.content:
                extracted_data['raw_text'] = result.content
            
            # Extrair tabelas estruturadas
            if result.tables:
                for table in result.tables:
                    table_data = []
                    for cell in table.cells:
                        table_data.append({
                            'row': cell.row_index,
                            'column': cell.column_index,
                            'content': cell.content,
                            'confidence': getattr(cell, 'confidence', 0.0)
                        })
                    extracted_data['tabelas'].append(table_data)
            
            # Extrair key-value pairs
            if hasattr(result, 'key_value_pairs') and result.key_value_pairs:
                for kv_pair in result.key_value_pairs:
                    key = kv_pair.key.content if kv_pair.key else ""
                    value = kv_pair.value.content if kv_pair.value else ""
                    self._map_key_value_to_fields(key, value, extracted_data)
            
            # Extrair dados específicos usando padrões
            self._extract_specific_patterns(extracted_data['raw_text'], extracted_data)
            
            # Calcular score de confiança
            extracted_data['confidence_score'] = self._calculate_confidence_score(extracted_data)
            
        except Exception as e:
            logger.error(f"Erro ao processar resultado Azure: {str(e)}")
        
        return extracted_data
    
    def _map_key_value_to_fields(self, key: str, value: str, data: Dict):
        """Mapeia key-value pairs para campos específicos"""
        key_lower = key.lower()
        
        if any(term in key_lower for term in ['empresa', 'razão social', 'nome']):
            data['empresa'] = value
        elif any(term in key_lower for term in ['cnpj', 'cpf']):
            cnpj = re.sub(r'[^\d]', '', value)
            if len(cnpj) >= 11:
                data['cnpj'] = value
        elif any(term in key_lower for term in ['metodologia', 'método', 'abordagem']):
            data['metodologia'] = value
        elif any(term in key_lower for term in ['prazo', 'cronograma', 'tempo', 'dias']):
            prazo = self._extract_number_from_text(value)
            if prazo > 0:
                data['prazo_dias'] = prazo
        elif any(term in key_lower for term in ['equipe', 'pessoas', 'profissionais']):
            equipe = self._extract_number_from_text(value)
            if equipe > 0:
                data['equipe_total'] = equipe
        elif any(term in key_lower for term in ['preço', 'valor', 'total']):
            preco = self._extract_currency_from_text(value)
            if preco > 0:
                data['preco_total'] = preco
        elif any(term in key_lower for term in ['bdi', 'benefício']):
            bdi = self._extract_percentage_from_text(value)
            if bdi > 0:
                data['bdi_percentual'] = bdi
    
    def _extract_specific_patterns(self, text: str, data: Dict):
        """Extrai padrões específicos do texto usando regex avançado"""
        
        # Padrões para empresa
        empresa_patterns = [
            r'(?:Empresa|Razão Social|Nome):\s*([A-Za-z\s]+(?:Ltda|S\.A\.|EIRELI)?)',
            r'([A-Za-z\s]+(?:Ltda|S\.A\.|EIRELI))',
        ]
        
        for pattern in empresa_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches and not data['empresa']:
                data['empresa'] = matches[0].strip()
                break
        
        # Padrões para CNPJ
        cnpj_patterns = [
            r'CNPJ[:\s]*(\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2})',
            r'(\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2})'
        ]
        
        for pattern in cnpj_patterns:
            matches = re.findall(pattern, text)
            if matches and not data['cnpj']:
                data['cnpj'] = matches[0]
                break
        
        # Padrões para metodologia (mais detalhados)
        metodologia_patterns = [
            r'(?:metodologia|abordagem)[:\s]*([^.]+(?:scrum|kanban|ágil|agile|cascata|waterfall)[^.]*)',
            r'((?:scrum|kanban|ágil|agile|cascata|waterfall)[^.]*)',
            r'(?:metodologia|método)[:\s]*([^.]{20,200})'
        ]
        
        for pattern in metodologia_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches and not data['metodologia']:
                data['metodologia'] = matches[0].strip()
                break
        
        # Extrair cronograma detalhado
        self._extract_cronograma_fases(text, data)
        
        # Extrair equipe detalhada
        self._extract_equipe_detalhada(text, data)
        
        # Padrões para prazo (melhorados)
        prazo_patterns = [
            r'(?:prazo|cronograma|tempo)[:\s]*(\d+)\s*dias?',
            r'(\d+)\s*dias?\s*(?:úteis|corridos)?',
            r'(?:em|dentro de)\s*(\d+)\s*dias?',
            r'(\d+)\s*semanas?'
        ]
        
        for pattern in prazo_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                prazo = int(matches[0])
                if 'semana' in pattern:
                    prazo *= 7
                if prazo > data['prazo_dias']:
                    data['prazo_dias'] = prazo
        
        # Padrões para equipe (melhorados)
        equipe_patterns = [
            r'(?:equipe|pessoas|profissionais)[:\s]*(\d+)',
            r'(\d+)\s*(?:pessoas|profissionais)',
            r'(?:composta por|formada por)\s*(\d+)'
        ]
        
        for pattern in equipe_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                equipe = int(matches[0])
                if equipe > data['equipe_total']:
                    data['equipe_total'] = equipe
        
        # Padrões para preço
        preco_patterns = [
            r'R\$\s*([\d.,]+)',
            r'(?:valor|preço|total)[:\s]*R\$\s*([\d.,]+)',
            r'([\d.,]+)\s*reais?'
        ]
        
        for pattern in preco_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                preco_str = matches[0].replace('.', '').replace(',', '.')
                try:
                    preco = float(preco_str)
                    if preco > data['preco_total']:
                        data['preco_total'] = preco
                except:
                    continue
        
        # Padrões para BDI
        bdi_patterns = [
            r'BDI[:\s]*(\d+(?:,\d+)?)\s*%',
            r'(?:benefício|lucro)[:\s]*(\d+(?:,\d+)?)\s*%'
        ]
        
        for pattern in bdi_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                bdi_str = matches[0].replace(',', '.')
                try:
                    bdi = float(bdi_str)
                    if bdi > data['bdi_percentual']:
                        data['bdi_percentual'] = bdi
                except:
                    continue
        
        # Extrair tecnologias
        tech_keywords = ['SAP', 'Microsoft', 'Oracle', 'Java', 'Python', 'SQL', 'Azure', 'AWS', 'Scrum', 'Kanban', 'React', 'Angular', 'Node.js']
        for tech in tech_keywords:
            if tech.lower() in text.lower():
                if tech not in data['tecnologias']:
                    data['tecnologias'].append(tech)
        
        # Extrair arquitetura do sistema
        self._extract_arquitetura_sistema(text, data)
        
        # Extrair requisitos técnicos
        self._extract_requisitos_tecnicos(text, data)
        
        # Extrair equipamentos e materiais das tabelas
        self._extract_items_from_tables(data)
    
    def _extract_cronograma_fases(self, text: str, data: Dict):
        """Extrai fases do cronograma"""
        fase_patterns = [
            r'(?:fase|etapa)\s*\d+[:\s]*([^.]{10,100})',
            r'(\d+[°º]?\s*(?:fase|etapa)[:\s]*[^.]{10,100})',
            r'((?:análise|desenvolvimento|teste|implantação)[^.]{10,100})'
        ]
        
        fases = []
        for pattern in fase_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches:
                if match not in fases and len(match.strip()) > 10:
                    fases.append(match.strip())
        
        data['cronograma_fases'] = fases[:6]  # Máximo 6 fases
    
    def _extract_equipe_detalhada(self, text: str, data: Dict):
        """Extrai detalhes da equipe"""
        equipe_patterns = [
            r'(\d+)\s*(?:gerente|coordenador|líder)',
            r'(\d+)\s*(?:desenvolvedor|programador)',
            r'(\d+)\s*(?:analista|arquiteto)',
            r'(\d+)\s*(?:testador|qa)'
        ]
        
        equipe_detalhada = {}
        for pattern in equipe_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                if 'gerente' in pattern or 'coordenador' in pattern or 'líder' in pattern:
                    equipe_detalhada['gerencia'] = int(matches[0])
                elif 'desenvolvedor' in pattern or 'programador' in pattern:
                    equipe_detalhada['desenvolvimento'] = int(matches[0])
                elif 'analista' in pattern or 'arquiteto' in pattern:
                    equipe_detalhada['analise'] = int(matches[0])
                elif 'testador' in pattern or 'qa' in pattern:
                    equipe_detalhada['testes'] = int(matches[0])
        
        data['equipe_detalhada'] = equipe_detalhada
    
    def _extract_arquitetura_sistema(self, text: str, data: Dict):
        """Extrai informações sobre arquitetura do sistema"""
        arquitetura_patterns = [
            r'(?:arquitetura|estrutura)[:\s]*([^.]{20,200})',
            r'(?:tecnologia|plataforma)[:\s]*([^.]{20,200})',
            r'(?:banco de dados|database)[:\s]*([^.]{10,100})'
        ]
        
        for pattern in arquitetura_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches and not data['arquitetura_sistema']:
                data['arquitetura_sistema'] = matches[0].strip()
                break
    
    def _extract_requisitos_tecnicos(self, text: str, data: Dict):
        """Extrai requisitos técnicos"""
        requisitos_keywords = [
            'alta disponibilidade', 'escalabilidade', 'performance', 'segurança',
            'backup', 'disaster recovery', 'load balancing', 'clustering',
            'ssl', 'https', 'criptografia', 'autenticação'
        ]
        
        requisitos = []
        for req in requisitos_keywords:
            if req.lower() in text.lower():
                requisitos.append(req.title())
        
        data['requisitos_tecnicos'] = requisitos
    
    def _extract_items_from_tables(self, data: Dict):
        """Extrai equipamentos e materiais das tabelas"""
        for table in data['tabelas']:
            for cell in table:
                content = cell['content'].lower()
                if any(term in content for term in ['servidor', 'computador', 'notebook', 'equipamento']):
                    if cell['content'] not in data['equipamentos']:
                        data['equipamentos'].append(cell['content'])
                elif any(term in content for term in ['licença', 'software', 'material']):
                    if cell['content'] not in data['materiais']:
                        data['materiais'].append(cell['content'])
    
    def _extract_number_from_text(self, text: str) -> int:
        """Extrai número de um texto"""
        numbers = re.findall(r'\d+', text)
        return int(numbers[0]) if numbers else 0
    
    def _extract_currency_from_text(self, text: str) -> float:
        """Extrai valor monetário de um texto"""
        # Remove símbolos e converte para float
        clean_text = re.sub(r'[^\d.,]', '', text)
        clean_text = clean_text.replace('.', '').replace(',', '.')
        try:
            return float(clean_text)
        except:
            return 0.0
    
    def _extract_percentage_from_text(self, text: str) -> float:
        """Extrai percentual de um texto"""
        numbers = re.findall(r'(\d+(?:,\d+)?)', text)
        if numbers:
            try:
                return float(numbers[0].replace(',', '.'))
            except:
                return 0.0
        return 0.0
    
    def _calculate_confidence_score(self, data: Dict) -> float:
        """Calcula score de confiança baseado nos dados extraídos"""
        score = 0.0
        total_fields = 10
        
        if data['empresa']:
            score += 1.0
        if data['cnpj']:
            score += 1.0
        if data['metodologia']:
            score += 1.0
        if data['prazo_dias'] > 0:
            score += 1.0
        if data['equipe_total'] > 0:
            score += 1.0
        if data['preco_total'] > 0:
            score += 1.0
        if data['bdi_percentual'] > 0:
            score += 1.0
        if data['tecnologias']:
            score += 1.0
        if data['cronograma_fases']:
            score += 1.0
        if data['arquitetura_sistema']:
            score += 1.0
        
        return (score / total_fields) * 100
    
    def _fallback_extraction(self, pdf_path: str) -> Dict[str, Any]:
        """Extração de fallback usando PyPDF2"""
        logger.warning("Usando extração de fallback")
        
        try:
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text()
            
            # Aplicar padrões básicos mesmo no fallback
            data = {
                'empresa': '',
                'cnpj': '',
                'metodologia': 'Metodologia não especificada',
                'metodologia_detalhada': '',
                'prazo_dias': 0,
                'cronograma_fases': [],
                'equipe_total': 0,
                'equipe_detalhada': {},
                'equipamentos': [],
                'materiais': [],
                'tecnologias': [],
                'arquitetura_sistema': '',
                'requisitos_tecnicos': [],
                'riscos_tecnicos': [],
                'preco_total': 0.0,
                'bdi_percentual': 0.0,
                'condicoes_pagamento': '',
                'garantia': '',
                'composicao_custos': {
                    'mao_obra': 0.0,
                    'materiais': 0.0,
                    'equipamentos': 0.0
                },
                'raw_text': text,
                'tabelas': [],
                'confidence_score': 25.0
            }
            
            self._extract_specific_patterns(text, data)
            return data
            
        except Exception as e:
            logger.error(f"Erro no fallback: {str(e)}")
            return {
                'empresa': 'Erro na extração',
                'cnpj': '',
                'metodologia': 'Erro na extração',
                'metodologia_detalhada': '',
                'prazo_dias': 0,
                'cronograma_fases': [],
                'equipe_total': 0,
                'equipe_detalhada': {},
                'equipamentos': [],
                'materiais': [],
                'tecnologias': [],
                'arquitetura_sistema': '',
                'requisitos_tecnicos': [],
                'riscos_tecnicos': [],
                'preco_total': 0.0,
                'bdi_percentual': 0.0,
                'condicoes_pagamento': '',
                'garantia': '',
                'composicao_custos': {
                    'mao_obra': 0.0,
                    'materiais': 0.0,
                    'equipamentos': 0.0
                },
                'raw_text': '',
                'tabelas': [],
                'confidence_score': 0.0
            }

class ExcelExtractor:
    """Extrator para arquivos Excel SEM PANDAS - usando openpyxl"""
    
    def extract_from_excel(self, excel_path: str) -> Dict[str, Any]:
        """Extrai dados comerciais do Excel usando openpyxl"""
        try:
            workbook = load_workbook(excel_path, read_only=True)
            
            extracted_data = {
                'empresa': '',
                'cnpj': '',
                'preco_total': 0.0,
                'bdi_percentual': 0.0,
                'condicoes_pagamento': '',
                'garantia': '',
                'composicao_custos': {
                    'mao_obra': 0.0,
                    'materiais': 0.0,
                    'equipamentos': 0.0
                }
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                if any(term in sheet_name.lower() for term in ['comercial', 'proposta']):
                    self._extract_commercial_data(sheet, extracted_data)
                elif any(term in sheet_name.lower() for term in ['custo', 'composição']):
                    self._extract_cost_composition(sheet, extracted_data)
            
            workbook.close()
            return extracted_data
            
        except Exception as e:
            logger.error(f"Erro ao extrair dados do Excel: {str(e)}")
            return {}
    
    def _extract_commercial_data(self, sheet, data: Dict):
        """Extrai dados comerciais da planilha"""
        for row in sheet.iter_rows(values_only=True):
            if not row or not any(row):
                continue
            
            row_text = ' '.join(str(cell) for cell in row if cell is not None).lower()
            
            # Buscar CNPJ
            cnpj_match = re.search(r'(\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2})', row_text)
            if cnpj_match and not data['cnpj']:
                data['cnpj'] = cnpj_match.group(1)
            
            # Buscar preço total
            if any(term in row_text for term in ['total', 'preço', 'valor']):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell > 1000:
                        data['preco_total'] = float(cell)
                    elif isinstance(cell, str) and 'R$' in str(cell):
                        price_match = re.search(r'R\$\s*([\d.,]+)', str(cell))
                        if price_match:
                            price_str = price_match.group(1).replace('.', '').replace(',', '.')
                            try:
                                data['preco_total'] = float(price_str)
                            except:
                                pass
            
            # Buscar BDI
            if 'bdi' in row_text:
                for cell in row:
                    if isinstance(cell, (int, float)) and 0 < cell < 100:
                        data['bdi_percentual'] = float(cell)
    
    def _extract_cost_composition(self, sheet, data: Dict):
        """Extrai composição de custos da planilha"""
        for row in sheet.iter_rows(values_only=True):
            if not row or not any(row):
                continue
            
            row_text = ' '.join(str(cell) for cell in row if cell is not None).lower()
            
            # Buscar valores por categoria
            value = None
            for cell in row:
                if isinstance(cell, (int, float)) and cell > 0:
                    value = float(cell)
                    break
            
            if value:
                if any(term in row_text for term in ['mão de obra', 'mao de obra', 'pessoal']):
                    data['composicao_custos']['mao_obra'] = value
                elif any(term in row_text for term in ['material', 'insumo']):
                    data['composicao_custos']['materiais'] = value
                elif any(term in row_text for term in ['equipamento', 'hardware']):
                    data['composicao_custos']['equipamentos'] = value



class ProposalAnalyzer:
    """Analisador de propostas SEM PANDAS - usando estruturas Python nativas"""
    
    def __init__(self):
        self.azure_extractor = AzureDocumentIntelligenceExtractor(AZURE_ENDPOINT, AZURE_KEY)
        self.excel_extractor = ExcelExtractor()
    
    def analyze_proposals(self, files: List[Dict]) -> Dict[str, Any]:
        """Analisa múltiplas propostas e gera comparação"""
        proposals = []
        
        for file_info in files:
            file_path = file_info['path']
            file_type = file_info['type']
            
            try:
                if file_type == 'pdf':
                    data = self.azure_extractor.extract_from_pdf(file_path)
                elif file_type in ['xlsx', 'xls']:
                    data = self.excel_extractor.extract_from_excel(file_path)
                else:
                    continue
                
                # Adicionar informações do arquivo
                data['filename'] = file_info['filename']
                data['file_type'] = file_type
                proposals.append(data)
                
            except Exception as e:
                logger.error(f"Erro ao processar {file_info['filename']}: {str(e)}")
                continue
        
        if not proposals:
            raise ValueError("Nenhuma proposta válida foi processada")
        
        # Consolidar dados (mesclar PDF + Excel da mesma empresa)
        consolidated_proposals = self._consolidate_proposals(proposals)
        
        # Calcular scores técnicos e comerciais
        self._calculate_technical_scores(consolidated_proposals)
        self._calculate_commercial_scores(consolidated_proposals)
        
        # Ordenar por ranking técnico
        consolidated_proposals.sort(key=lambda x: x.get('score_tecnico', 0), reverse=True)
        
        return {
            'proposals': consolidated_proposals,
            'summary': self._generate_summary(consolidated_proposals),
            'analysis_date': datetime.now().strftime('%d/%m/%Y às %H:%M')
        }
    
    def _consolidate_proposals(self, proposals: List[Dict]) -> List[Dict]:
        """Consolida dados de PDF e Excel da mesma empresa"""
        consolidated = {}
        
        for proposal in proposals:
            empresa = proposal.get('empresa', '').strip()
            
            # Tentar identificar empresa por CNPJ se nome não estiver disponível
            if not empresa:
                cnpj = proposal.get('cnpj', '')
                if cnpj:
                    empresa = f"Empresa {cnpj[:8]}"
                else:
                    empresa = f"Empresa {len(consolidated) + 1}"
            
            if empresa not in consolidated:
                consolidated[empresa] = {
                    'empresa': empresa,
                    'cnpj': '',
                    'metodologia': '',
                    'metodologia_detalhada': '',
                    'prazo_dias': 0,
                    'cronograma_fases': [],
                    'equipe_total': 0,
                    'equipe_detalhada': {},
                    'equipamentos': [],
                    'materiais': [],
                    'tecnologias': [],
                    'arquitetura_sistema': '',
                    'requisitos_tecnicos': [],
                    'riscos_tecnicos': [],
                    'preco_total': 0.0,
                    'bdi_percentual': 0.0,
                    'condicoes_pagamento': '',
                    'garantia': '',
                    'composicao_custos': {
                        'mao_obra': 0.0,
                        'materiais': 0.0,
                        'equipamentos': 0.0
                    },
                    'confidence_score': 0.0,
                    'score_tecnico': 0.0,
                    'score_comercial': 0.0,
                    'files_processed': []
                }
            
            # Mesclar dados
            current = consolidated[empresa]
            
            # Atualizar campos se não estiverem preenchidos ou se o novo valor for melhor
            if proposal.get('cnpj') and not current['cnpj']:
                current['cnpj'] = proposal['cnpj']
            
            if proposal.get('metodologia') and not current['metodologia']:
                current['metodologia'] = proposal['metodologia']
            
            if proposal.get('metodologia_detalhada') and not current['metodologia_detalhada']:
                current['metodologia_detalhada'] = proposal['metodologia_detalhada']
            
            if proposal.get('prazo_dias', 0) > current['prazo_dias']:
                current['prazo_dias'] = proposal['prazo_dias']
            
            if proposal.get('equipe_total', 0) > current['equipe_total']:
                current['equipe_total'] = proposal['equipe_total']
            
            if proposal.get('preco_total', 0) > current['preco_total']:
                current['preco_total'] = proposal['preco_total']
            
            if proposal.get('bdi_percentual', 0) > current['bdi_percentual']:
                current['bdi_percentual'] = proposal['bdi_percentual']
            
            if proposal.get('arquitetura_sistema') and not current['arquitetura_sistema']:
                current['arquitetura_sistema'] = proposal['arquitetura_sistema']
            
            # Mesclar listas
            for item in proposal.get('cronograma_fases', []):
                if item not in current['cronograma_fases']:
                    current['cronograma_fases'].append(item)
            
            for item in proposal.get('equipamentos', []):
                if item not in current['equipamentos']:
                    current['equipamentos'].append(item)
            
            for item in proposal.get('materiais', []):
                if item not in current['materiais']:
                    current['materiais'].append(item)
            
            for item in proposal.get('tecnologias', []):
                if item not in current['tecnologias']:
                    current['tecnologias'].append(item)
            
            for item in proposal.get('requisitos_tecnicos', []):
                if item not in current['requisitos_tecnicos']:
                    current['requisitos_tecnicos'].append(item)
            
            # Mesclar dicionários
            for key, value in proposal.get('equipe_detalhada', {}).items():
                if value > current['equipe_detalhada'].get(key, 0):
                    current['equipe_detalhada'][key] = value
            
            # Atualizar composição de custos
            for key, value in proposal.get('composicao_custos', {}).items():
                if value > current['composicao_custos'].get(key, 0):
                    current['composicao_custos'][key] = value
            
            # Atualizar confidence score (usar o maior)
            if proposal.get('confidence_score', 0) > current['confidence_score']:
                current['confidence_score'] = proposal['confidence_score']
            
            # Adicionar arquivo processado
            current['files_processed'].append(proposal.get('filename', 'unknown'))
        
        return list(consolidated.values())
    
    def _calculate_technical_scores(self, proposals: List[Dict]):
        """Calcula scores técnicos baseado nos dados extraídos"""
        for proposal in proposals:
            score = 0.0
            max_score = 100.0
            
            # Metodologia (30 pontos)
            if proposal['metodologia'] and proposal['metodologia'] != 'Metodologia não especificada':
                metodologia_score = 20.0
                # Bonus para metodologias ágeis
                if any(term in proposal['metodologia'].lower() for term in ['scrum', 'kanban', 'ágil', 'agile']):
                    metodologia_score = 30.0
                elif any(term in proposal['metodologia'].lower() for term in ['cascata', 'waterfall']):
                    metodologia_score = 20.0
                else:
                    metodologia_score = 25.0
                score += metodologia_score
            
            # Prazo (25 pontos)
            if proposal['prazo_dias'] > 0:
                if proposal['prazo_dias'] <= 90:
                    score += 25.0  # Prazo excelente
                elif proposal['prazo_dias'] <= 120:
                    score += 20.0  # Prazo bom
                elif proposal['prazo_dias'] <= 150:
                    score += 15.0  # Prazo aceitável
                else:
                    score += 10.0   # Prazo ruim
            
            # Equipe (20 pontos)
            if proposal['equipe_total'] > 0:
                if proposal['equipe_total'] >= 8:
                    score += 20.0  # Equipe robusta
                elif proposal['equipe_total'] >= 5:
                    score += 15.0  # Equipe adequada
                elif proposal['equipe_total'] >= 3:
                    score += 10.0  # Equipe mínima
                else:
                    score += 5.0   # Equipe insuficiente
            
            # Cronograma detalhado (10 pontos)
            if proposal['cronograma_fases']:
                score += min(len(proposal['cronograma_fases']) * 2, 10)
            
            # Tecnologias (10 pontos)
            if proposal['tecnologias']:
                tech_score = min(len(proposal['tecnologias']) * 2, 10)
                score += tech_score
            
            # Arquitetura e requisitos técnicos (15 pontos)
            if proposal['arquitetura_sistema']:
                score += 7.5
            if proposal['requisitos_tecnicos']:
                score += min(len(proposal['requisitos_tecnicos']) * 1.5, 7.5)
            
            proposal['score_tecnico'] = round(score, 1)
    
    def _calculate_commercial_scores(self, proposals: List[Dict]):
        """Calcula scores comerciais baseado nos dados extraídos"""
        # Filtrar propostas com preço
        proposals_with_price = [p for p in proposals if p['preco_total'] > 0]
        
        if not proposals_with_price:
            for proposal in proposals:
                proposal['score_comercial'] = 0.0
            return
        
        # Ordenar por preço para calcular ranking
        proposals_with_price.sort(key=lambda x: x['preco_total'])
        
        for i, proposal in enumerate(proposals_with_price):
            score = 0.0
            
            # Ranking de preço (50 pontos)
            if i == 0:
                score += 50.0  # Melhor preço
            elif i == 1:
                score += 40.0  # Segundo melhor
            elif i == 2:
                score += 30.0  # Terceiro melhor
            else:
                score += 20.0  # Demais
            
            # BDI razoável (20 pontos)
            if proposal['bdi_percentual'] > 0:
                if proposal['bdi_percentual'] <= 25:
                    score += 20.0  # BDI excelente
                elif proposal['bdi_percentual'] <= 35:
                    score += 15.0  # BDI bom
                elif proposal['bdi_percentual'] <= 45:
                    score += 10.0  # BDI aceitável
                else:
                    score += 5.0   # BDI alto
            
            # Composição de custos detalhada (15 pontos)
            custos = proposal['composicao_custos']
            if any(custos.values()):
                score += 15.0
            
            # Condições comerciais (15 pontos)
            if proposal['condicoes_pagamento']:
                score += 7.5
            if proposal['garantia']:
                score += 7.5
            
            proposal['score_comercial'] = round(score, 1)
        
        # Zerar score das propostas sem preço
        for proposal in proposals:
            if proposal['preco_total'] == 0:
                proposal['score_comercial'] = 0.0
    
    def _generate_summary(self, proposals: List[Dict]) -> Dict[str, Any]:
        """Gera resumo da análise"""
        if not proposals:
            return {}
        
        # Encontrar melhor proposta técnica e comercial
        best_technical = max(proposals, key=lambda x: x['score_tecnico'])
        
        proposals_with_price = [p for p in proposals if p['preco_total'] > 0]
        best_commercial = max(proposals_with_price, key=lambda x: x['score_comercial']) if proposals_with_price else None
        
        # Calcular estatísticas
        precos = [p['preco_total'] for p in proposals if p['preco_total'] > 0]
        prazos = [p['prazo_dias'] for p in proposals if p['prazo_dias'] > 0]
        
        return {
            'total_proposals': len(proposals),
            'best_technical': best_technical['empresa'] if best_technical else '',
            'best_commercial': best_commercial['empresa'] if best_commercial else '',
            'price_range': {
                'min': min(precos) if precos else 0,
                'max': max(precos) if precos else 0,
                'avg': sum(precos) / len(precos) if precos else 0
            },
            'deadline_range': {
                'min': min(prazos) if prazos else 0,
                'max': max(prazos) if prazos else 0,
                'avg': sum(prazos) / len(prazos) if prazos else 0
            }
        }

class TechnicalReportGenerator:
    """Gerador de relatório técnico especializado"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Configura estilos personalizados para o relatório técnico"""
        # Título principal
        self.styles.add(ParagraphStyle(
            name='TechnicalTitle',
            parent=self.styles['Title'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue,
            fontName='Helvetica-Bold'
        ))
        
        # Subtítulo técnico
        self.styles.add(ParagraphStyle(
            name='TechnicalSubtitle',
            parent=self.styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            spaceBefore=20,
            textColor=colors.darkblue,
            fontName='Helvetica-Bold'
        ))
        
        # Cabeçalho de seção técnica
        self.styles.add(ParagraphStyle(
            name='TechnicalSectionHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=15,
            spaceBefore=15,
            textColor=colors.darkgreen,
            fontName='Helvetica-Bold',
            borderWidth=1,
            borderColor=colors.darkgreen,
            borderPadding=5
        ))
        
        # Texto normal técnico
        self.styles.add(ParagraphStyle(
            name='TechnicalNormal',
            parent=self.styles['Normal'],
            fontSize=11,
            spaceAfter=10,
            fontName='Helvetica'
        ))
    
    def generate_technical_report(self, analysis_result: Dict[str, Any], output_path: str):
        """Gera relatório técnico especializado"""
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        story = []
        
        # Cabeçalho do relatório técnico
        self._add_technical_header(story, analysis_result)
        
        # Seção 1: Resumo Técnico do TR
        self._add_technical_tr_summary(story)
        
        # Seção 2: Matriz de Comparação Técnica
        self._add_technical_matrix(story, analysis_result['proposals'])
        
        # Seção 3: Ranking e Scores Técnicos
        self._add_technical_ranking(story, analysis_result['proposals'])
        
        # Seção 4: Análise Detalhada por Empresa
        self._add_detailed_technical_analysis(story, analysis_result['proposals'])
        
        # Seção 5: Recomendações Técnicas
        self._add_technical_recommendations(story, analysis_result)
        
        doc.build(story)
        logger.info(f"Relatório técnico gerado: {output_path}")
    
    def _add_technical_header(self, story: List, analysis_result: Dict):
        """Adiciona cabeçalho do relatório técnico"""
        # Título principal
        title = Paragraph("ANÁLISE E EQUALIZAÇÃO TÉCNICA DE PROPOSTAS", self.styles['TechnicalTitle'])
        story.append(title)
        
        # Subtítulo
        subtitle = Paragraph("Avaliação Técnica Especializada", self.styles['TechnicalSubtitle'])
        story.append(subtitle)
        
        # Data de geração
        date_text = f"<b>Data de Geração:</b> {analysis_result['analysis_date']}"
        date_para = Paragraph(date_text, self.styles['TechnicalNormal'])
        story.append(date_para)
        
        # Linha separadora
        story.append(Spacer(1, 20))
        story.append(self._create_separator_line())
        story.append(Spacer(1, 20))
    
    def _add_technical_tr_summary(self, story: List):
        """Adiciona resumo técnico do TR"""
        section_title = Paragraph("🔧 SEÇÃO 1: RESUMO TÉCNICO DO TERMO DE REFERÊNCIA", self.styles['TechnicalSectionHeader'])
        story.append(section_title)
        
        # Objeto técnico
        story.append(Paragraph("<b>Objeto Técnico</b>", self.styles['TechnicalSubtitle']))
        story.append(Paragraph("Sistema de Gestão Empresarial Integrado", self.styles['TechnicalNormal']))
        
        # Especificações técnicas detalhadas
        story.append(Paragraph("<b>Especificações Técnicas Obrigatórias</b>", self.styles['TechnicalSubtitle']))
        specs = [
            "• <b>Arquitetura:</b> Sistema integrado com módulos interoperáveis",
            "• <b>Módulos Funcionais:</b> Financeiro, Estoque, Vendas, Compras, RH",
            "• <b>Interface:</b> Web responsiva com suporte a dispositivos móveis",
            "• <b>Banco de Dados:</b> Robusto, escalável e com backup automático",
            "• <b>Relatórios:</b> Gerenciais customizáveis e dashboards em tempo real",
            "• <b>Segurança:</b> Autenticação, autorização e criptografia de dados",
            "• <b>Performance:</b> Suporte a múltiplos usuários simultâneos"
        ]
        for spec in specs:
            story.append(Paragraph(spec, self.styles['TechnicalNormal']))
        
        # Metodologia técnica exigida
        story.append(Paragraph("<b>Metodologia de Desenvolvimento Exigida</b>", self.styles['TechnicalSubtitle']))
        metodologia = [
            "• <b>Abordagem:</b> Metodologia ágil (Scrum/Kanban) ou híbrida",
            "• <b>Fases Obrigatórias:</b>",
            "  - Levantamento e análise de requisitos",
            "  - Design e arquitetura do sistema",
            "  - Desenvolvimento iterativo",
            "  - Testes integrados e validação",
            "  - Implantação e go-live",
            "  - Suporte e manutenção",
            "• <b>Documentação:</b> Técnica completa e manuais de usuário",
            "• <b>Treinamento:</b> Equipe técnica e usuários finais"
        ]
        for item in metodologia:
            story.append(Paragraph(item, self.styles['TechnicalNormal']))
        
        # Critérios técnicos de avaliação
        story.append(Paragraph("<b>Critérios de Avaliação Técnica</b>", self.styles['TechnicalSubtitle']))
        story.append(Paragraph("• <b>Peso na Avaliação:</b> 70% da nota final", self.styles['TechnicalNormal']))
        story.append(Paragraph("• <b>Metodologia:</b> 30% da nota técnica", self.styles['TechnicalNormal']))
        story.append(Paragraph("• <b>Cronograma:</b> 25% da nota técnica", self.styles['TechnicalNormal']))
        story.append(Paragraph("• <b>Equipe Técnica:</b> 20% da nota técnica", self.styles['TechnicalNormal']))
        story.append(Paragraph("• <b>Arquitetura/Tecnologia:</b> 15% da nota técnica", self.styles['TechnicalNormal']))
        story.append(Paragraph("• <b>Recursos e Ferramentas:</b> 10% da nota técnica", self.styles['TechnicalNormal']))
        
        story.append(self._create_separator_line())
    
    def _add_technical_matrix(self, story: List, proposals: List[Dict]):
        """Adiciona matriz de comparação técnica"""
        section_title = Paragraph("📊 SEÇÃO 2: MATRIZ DE COMPARAÇÃO TÉCNICA", self.styles['TechnicalSectionHeader'])
        story.append(section_title)
        
        # Criar tabela de comparação técnica
        table_data = [['Empresa', 'Metodologia', 'Prazo', 'Equipe', 'Tecnologias', 'Arquitetura', 'Score Técnico']]
        
        for proposal in proposals:
            metodologia_check = "✓" if proposal['metodologia'] and proposal['metodologia'] != 'Metodologia não especificada' else "✗"
            prazo_check = "✓" if proposal['prazo_dias'] > 0 and proposal['prazo_dias'] <= 120 else "✗"
            equipe_check = "✓" if proposal['equipe_total'] >= 5 else "✗"
            tech_check = "✓" if proposal['tecnologias'] else "✗"
            arq_check = "✓" if proposal['arquitetura_sistema'] else "✗"
            
            table_data.append([
                f"<b>{proposal['empresa']}</b>",
                metodologia_check,
                prazo_check,
                equipe_check,
                tech_check,
                arq_check,
                f"<b>{proposal['score_tecnico']:.1f}%</b>"
            ])
        
        table = Table(table_data, colWidths=[3.5*cm, 2*cm, 1.5*cm, 1.5*cm, 2*cm, 2*cm, 2*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
        
        story.append(self._create_separator_line())
    
    def _add_technical_ranking(self, story: List, proposals: List[Dict]):
        """Adiciona ranking técnico"""
        section_title = Paragraph("🏆 SEÇÃO 3: RANKING E SCORES TÉCNICOS", self.styles['TechnicalSectionHeader'])
        story.append(section_title)
        
        # Ranking técnico final
        story.append(Paragraph("🥇 Ranking Técnico Final", self.styles['TechnicalSubtitle']))
        
        for i, proposal in enumerate(proposals, 1):
            if i == 1:
                icon = "🥇"
            elif i == 2:
                icon = "🥈"
            elif i == 3:
                icon = "🥉"
            else:
                icon = f"{i}º"
            
            ranking_text = f"{icon} <b>{proposal['empresa']}</b> - {proposal['score_tecnico']:.1f}%"
            story.append(Paragraph(ranking_text, self.styles['TechnicalNormal']))
        
        story.append(Spacer(1, 20))
        
        # Análise de scores
        story.append(Paragraph("📈 Análise de Scores Técnicos", self.styles['TechnicalSubtitle']))
        
        if proposals:
            best_score = proposals[0]['score_tecnico']
            worst_score = min(p['score_tecnico'] for p in proposals)
            avg_score = sum(p['score_tecnico'] for p in proposals) / len(proposals)
            
            story.append(Paragraph(f"• <b>Melhor Score:</b> {best_score:.1f}%", self.styles['TechnicalNormal']))
            story.append(Paragraph(f"• <b>Pior Score:</b> {worst_score:.1f}%", self.styles['TechnicalNormal']))
            story.append(Paragraph(f"• <b>Score Médio:</b> {avg_score:.1f}%", self.styles['TechnicalNormal']))
            story.append(Paragraph(f"• <b>Diferença:</b> {best_score - worst_score:.1f} pontos", self.styles['TechnicalNormal']))
        
        story.append(self._create_separator_line())
    
    def _add_detailed_technical_analysis(self, story: List, proposals: List[Dict]):
        """Adiciona análise técnica detalhada por empresa"""
        section_title = Paragraph("🔍 SEÇÃO 4: ANÁLISE TÉCNICA DETALHADA POR EMPRESA", self.styles['TechnicalSectionHeader'])
        story.append(section_title)
        
        for proposal in proposals:
            self._add_company_detailed_technical_analysis(story, proposal)
        
        story.append(self._create_separator_line())
    
    def _add_company_detailed_technical_analysis(self, story: List, proposal: Dict):
        """Adiciona análise técnica detalhada de uma empresa"""
        # Nome da empresa com score
        company_title = f"🏢 {proposal['empresa']} - Score: {proposal['score_tecnico']:.1f}%"
        story.append(Paragraph(company_title, self.styles['TechnicalSubtitle']))
        
        # Metodologia detalhada
        story.append(Paragraph("📋 Metodologia de Desenvolvimento:", self.styles['TechnicalNormal']))
        metodologia = proposal['metodologia'] if proposal['metodologia'] else "Não especificada"
        story.append(Paragraph(f"• <b>Abordagem:</b> {metodologia}", self.styles['TechnicalNormal']))
        
        if proposal['cronograma_fases']:
            story.append(Paragraph("• <b>Fases do Cronograma:</b>", self.styles['TechnicalNormal']))
            for fase in proposal['cronograma_fases']:
                story.append(Paragraph(f"  - {fase}", self.styles['TechnicalNormal']))
        
        aderencia = "✓ Adequada" if proposal['metodologia'] and proposal['metodologia'] != 'Metodologia não especificada' else "✗ Não especificada"
        story.append(Paragraph(f"• <b>Aderência ao TR:</b> {aderencia}", self.styles['TechnicalNormal']))
        
        # Cronograma e prazo
        story.append(Paragraph("⏰ Cronograma e Prazo:", self.styles['TechnicalNormal']))
        prazo = f"{proposal['prazo_dias']} dias" if proposal['prazo_dias'] > 0 else "Não informado"
        story.append(Paragraph(f"• <b>Prazo Total:</b> {prazo}", self.styles['TechnicalNormal']))
        
        if proposal['prazo_dias'] > 0:
            if proposal['prazo_dias'] <= 90:
                viabilidade = "✓ Excelente (≤ 90 dias)"
            elif proposal['prazo_dias'] <= 120:
                viabilidade = "✓ Dentro do prazo (≤ 120 dias)"
            else:
                viabilidade = "⚠️ Acima do prazo máximo"
        else:
            viabilidade = "✗ Não informado"
        story.append(Paragraph(f"• <b>Viabilidade:</b> {viabilidade}", self.styles['TechnicalNormal']))
        
        # Equipe técnica
        story.append(Paragraph("👥 Equipe Técnica:", self.styles['TechnicalNormal']))
        equipe = f"{proposal['equipe_total']} pessoas" if proposal['equipe_total'] > 0 else "Não informada"
        story.append(Paragraph(f"• <b>Total:</b> {equipe}", self.styles['TechnicalNormal']))
        
        if proposal['equipe_detalhada']:
            story.append(Paragraph("• <b>Composição da Equipe:</b>", self.styles['TechnicalNormal']))
            for cargo, qtd in proposal['equipe_detalhada'].items():
                story.append(Paragraph(f"  - {cargo.title()}: {qtd} pessoa(s)", self.styles['TechnicalNormal']))
        
        if proposal['equipe_total'] >= 8:
            status_equipe = "✓ Equipe robusta"
        elif proposal['equipe_total'] >= 5:
            status_equipe = "✓ Equipe adequada"
        elif proposal['equipe_total'] >= 3:
            status_equipe = "⚠️ Equipe mínima"
        else:
            status_equipe = "✗ Equipe insuficiente ou não informada"
        story.append(Paragraph(f"• <b>Avaliação:</b> {status_equipe}", self.styles['TechnicalNormal']))
        
        # Arquitetura e tecnologias
        story.append(Paragraph("🏗️ Arquitetura e Tecnologias:", self.styles['TechnicalNormal']))
        
        if proposal['arquitetura_sistema']:
            story.append(Paragraph(f"• <b>Arquitetura:</b> {proposal['arquitetura_sistema']}", self.styles['TechnicalNormal']))
        else:
            story.append(Paragraph("• <b>Arquitetura:</b> Não especificada", self.styles['TechnicalNormal']))
        
        if proposal['tecnologias']:
            tech_list = ", ".join(proposal['tecnologias'])
            story.append(Paragraph(f"• <b>Tecnologias:</b> {tech_list}", self.styles['TechnicalNormal']))
        else:
            story.append(Paragraph("• <b>Tecnologias:</b> Não especificadas", self.styles['TechnicalNormal']))
        
        if proposal['requisitos_tecnicos']:
            req_list = ", ".join(proposal['requisitos_tecnicos'])
            story.append(Paragraph(f"• <b>Requisitos Técnicos:</b> {req_list}", self.styles['TechnicalNormal']))
        
        # Recursos técnicos
        story.append(Paragraph("🔧 Recursos Técnicos:", self.styles['TechnicalNormal']))
        equipamentos_count = len(proposal['equipamentos'])
        materiais_count = len(proposal['materiais'])
        story.append(Paragraph(f"• <b>Equipamentos:</b> {equipamentos_count} itens listados", self.styles['TechnicalNormal']))
        story.append(Paragraph(f"• <b>Materiais/Software:</b> {materiais_count} itens listados", self.styles['TechnicalNormal']))
        
        # Pontos fortes técnicos
        pontos_fortes = []
        if proposal['metodologia'] and proposal['metodologia'] != 'Metodologia não especificada':
            if any(term in proposal['metodologia'].lower() for term in ['scrum', 'kanban', 'ágil', 'agile']):
                pontos_fortes.append("Metodologia ágil moderna")
            else:
                pontos_fortes.append("Metodologia bem definida")
        
        if proposal['prazo_dias'] > 0 and proposal['prazo_dias'] <= 90:
            pontos_fortes.append("Prazo otimizado")
        
        if proposal['equipe_total'] >= 8:
            pontos_fortes.append("Equipe técnica robusta")
        
        if len(proposal['tecnologias']) >= 3:
            pontos_fortes.append("Stack tecnológico diversificado")
        
        if proposal['arquitetura_sistema']:
            pontos_fortes.append("Arquitetura bem especificada")
        
        if pontos_fortes:
            story.append(Paragraph("✅ Pontos Fortes Técnicos:", self.styles['TechnicalNormal']))
            for ponto in pontos_fortes:
                story.append(Paragraph(f"• {ponto}", self.styles['TechnicalNormal']))
        
        # Gaps e riscos técnicos
        gaps = []
        if not proposal['metodologia'] or proposal['metodologia'] == 'Metodologia não especificada':
            gaps.append("Metodologia não especificada")
        
        if proposal['prazo_dias'] == 0:
            gaps.append("Cronograma não detalhado")
        elif proposal['prazo_dias'] > 120:
            gaps.append("Prazo acima do limite")
        
        if proposal['equipe_total'] < 5:
            gaps.append("Equipe técnica insuficiente")
        
        if not proposal['tecnologias']:
            gaps.append("Stack tecnológico não especificado")
        
        if not proposal['arquitetura_sistema']:
            gaps.append("Arquitetura do sistema não detalhada")
        
        if gaps:
            story.append(Paragraph("⚠️ Gaps e Riscos Técnicos:", self.styles['TechnicalNormal']))
            for gap in gaps:
                story.append(Paragraph(f"• {gap}", self.styles['TechnicalNormal']))
        
        story.append(Spacer(1, 15))
    
    def _add_technical_recommendations(self, story: List, analysis_result: Dict):
        """Adiciona recomendações técnicas"""
        section_title = Paragraph("💡 SEÇÃO 5: RECOMENDAÇÕES TÉCNICAS", self.styles['TechnicalSectionHeader'])
        story.append(section_title)
        
        proposals = analysis_result['proposals']
        
        if proposals:
            best_technical = proposals[0]  # Já ordenado por score técnico
            
            # Recomendação principal
            story.append(Paragraph("🏆 Recomendação Técnica Principal", self.styles['TechnicalSubtitle']))
            story.append(Paragraph(f"Com base na análise técnica detalhada, recomenda-se a empresa <b>{best_technical['empresa']}</b> que obteve o melhor score técnico ({best_technical['score_tecnico']:.1f}%).", self.styles['TechnicalNormal']))
            
            # Justificativa técnica
            story.append(Paragraph("📋 Justificativa Técnica:", self.styles['TechnicalSubtitle']))
            justificativas = []
            
            if best_technical['metodologia'] and best_technical['metodologia'] != 'Metodologia não especificada':
                justificativas.append(f"Metodologia bem definida: {best_technical['metodologia']}")
            
            if best_technical['prazo_dias'] > 0:
                justificativas.append(f"Cronograma viável: {best_technical['prazo_dias']} dias")
            
            if best_technical['equipe_total'] > 0:
                justificativas.append(f"Equipe adequada: {best_technical['equipe_total']} profissionais")
            
            if best_technical['tecnologias']:
                justificativas.append(f"Stack tecnológico: {', '.join(best_technical['tecnologias'][:3])}")
            
            for justificativa in justificativas:
                story.append(Paragraph(f"• {justificativa}", self.styles['TechnicalNormal']))
        
        # Ações técnicas recomendadas
        story.append(Paragraph("🔧 Ações Técnicas Recomendadas", self.styles['TechnicalSubtitle']))
        actions = [
            "• Solicitar detalhamento da arquitetura do sistema às empresas finalistas",
            "• Validar experiência da equipe técnica proposta em projetos similares",
            "• Confirmar disponibilidade dos profissionais para o período do projeto",
            "• Solicitar cronograma detalhado com marcos e entregas",
            "• Avaliar infraestrutura tecnológica necessária",
            "• Definir critérios de aceite para cada fase do desenvolvimento",
            "• Estabelecer métricas de qualidade e performance",
            "• Planejar estratégia de testes e homologação"
        ]
        
        for action in actions:
            story.append(Paragraph(action, self.styles['TechnicalNormal']))
        
        # Próximos passos técnicos
        story.append(Paragraph("🚀 Próximos Passos Técnicos", self.styles['TechnicalSubtitle']))
        next_steps = [
            "• Reunião técnica com as empresas finalistas",
            "• Apresentação da arquitetura proposta",
            "• Demonstração de cases similares",
            "• Validação de referências técnicas",
            "• Definição de ambiente de desenvolvimento",
            "• Elaboração do plano de projeto detalhado"
        ]
        
        for step in next_steps:
            story.append(Paragraph(step, self.styles['TechnicalNormal']))
    
    def _create_separator_line(self):
        """Cria linha separadora"""
        return Table([['---']], colWidths=[15*cm], style=TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 14),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.grey)
        ]))

class CommercialReportGenerator:
    """Gerador de relatório comercial especializado"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Configura estilos personalizados para o relatório comercial"""
        # Título principal
        self.styles.add(ParagraphStyle(
            name='CommercialTitle',
            parent=self.styles['Title'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkgreen,
            fontName='Helvetica-Bold'
        ))
        
        # Subtítulo comercial
        self.styles.add(ParagraphStyle(
            name='CommercialSubtitle',
            parent=self.styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            spaceBefore=20,
            textColor=colors.darkgreen,
            fontName='Helvetica-Bold'
        ))
        
        # Cabeçalho de seção comercial
        self.styles.add(ParagraphStyle(
            name='CommercialSectionHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=15,
            spaceBefore=15,
            textColor=colors.darkred,
            fontName='Helvetica-Bold',
            borderWidth=1,
            borderColor=colors.darkred,
            borderPadding=5
        ))
        
        # Texto normal comercial
        self.styles.add(ParagraphStyle(
            name='CommercialNormal',
            parent=self.styles['Normal'],
            fontSize=11,
            spaceAfter=10,
            fontName='Helvetica'
        ))
    
    def generate_commercial_report(self, analysis_result: Dict[str, Any], output_path: str):
        """Gera relatório comercial especializado"""
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        story = []
        
        # Cabeçalho do relatório comercial
        self._add_commercial_header(story, analysis_result)
        
        # Seção 1: Resumo Comercial do TR
        self._add_commercial_tr_summary(story)
        
        # Seção 2: Ranking de Preços
        self._add_price_ranking(story, analysis_result['proposals'])
        
        # Seção 3: Análise de Custos e BDI
        self._add_cost_analysis(story, analysis_result['proposals'])
        
        # Seção 4: Condições Comerciais
        self._add_commercial_conditions(story, analysis_result['proposals'])
        
        # Seção 5: Recomendações Comerciais
        self._add_commercial_recommendations(story, analysis_result)
        
        doc.build(story)
        logger.info(f"Relatório comercial gerado: {output_path}")
    
    def _add_commercial_header(self, story: List, analysis_result: Dict):
        """Adiciona cabeçalho do relatório comercial"""
        # Título principal
        title = Paragraph("ANÁLISE E EQUALIZAÇÃO COMERCIAL DE PROPOSTAS", self.styles['CommercialTitle'])
        story.append(title)
        
        # Subtítulo
        subtitle = Paragraph("Avaliação Comercial e Financeira", self.styles['CommercialSubtitle'])
        story.append(subtitle)
        
        # Data de geração
        date_text = f"<b>Data de Geração:</b> {analysis_result['analysis_date']}"
        date_para = Paragraph(date_text, self.styles['CommercialNormal'])
        story.append(date_para)
        
        # Linha separadora
        story.append(Spacer(1, 20))
        story.append(self._create_separator_line())
        story.append(Spacer(1, 20))
    
    def _add_commercial_tr_summary(self, story: List):
        """Adiciona resumo comercial do TR"""
        section_title = Paragraph("💰 SEÇÃO 1: RESUMO COMERCIAL DO TERMO DE REFERÊNCIA", self.styles['CommercialSectionHeader'])
        story.append(section_title)
        
        # Objeto comercial
        story.append(Paragraph("<b>Objeto da Contratação</b>", self.styles['CommercialSubtitle']))
        story.append(Paragraph("Desenvolvimento e implantação de Sistema de Gestão Empresarial", self.styles['CommercialNormal']))
        
        # Critérios comerciais
        story.append(Paragraph("<b>Critérios de Avaliação Comercial</b>", self.styles['CommercialSubtitle']))
        criterios = [
            "• <b>Peso na Avaliação:</b> 30% da nota final",
            "• <b>Menor Preço:</b> 50% da nota comercial",
            "• <b>BDI Adequado:</b> 20% da nota comercial",
            "• <b>Condições de Pagamento:</b> 15% da nota comercial",
            "• <b>Garantias Oferecidas:</b> 15% da nota comercial"
        ]
        for criterio in criterios:
            story.append(Paragraph(criterio, self.styles['CommercialNormal']))
        
        # Condições comerciais exigidas
        story.append(Paragraph("<b>Condições Comerciais Obrigatórias</b>", self.styles['CommercialSubtitle']))
        condicoes = [
            "• <b>Forma de Pagamento:</b> Conforme cronograma de entregas",
            "• <b>Garantia Mínima:</b> 12 meses para o sistema",
            "• <b>BDI Máximo:</b> 45% sobre custos diretos",
            "• <b>Reajuste:</b> Anual pelo IPCA",
            "• <b>Multas:</b> Por atraso na entrega",
            "• <b>Vigência:</b> 12 meses + renovações"
        ]
        for condicao in condicoes:
            story.append(Paragraph(condicao, self.styles['CommercialNormal']))
        
        story.append(self._create_separator_line())
    
    def _add_price_ranking(self, story: List, proposals: List[Dict]):
        """Adiciona ranking de preços"""
        section_title = Paragraph("📊 SEÇÃO 2: RANKING DE PREÇOS", self.styles['CommercialSectionHeader'])
        story.append(section_title)
        
        # Filtrar propostas com preço
        proposals_with_price = [p for p in proposals if p['preco_total'] > 0]
        
        if not proposals_with_price:
            story.append(Paragraph("⚠️ Nenhuma proposta com informações comerciais válidas foi encontrada.", self.styles['CommercialNormal']))
            return
        
        # Ordenar por preço
        proposals_with_price.sort(key=lambda x: x['preco_total'])
        
        # Tabela de ranking
        story.append(Paragraph("💵 Ranking por Menor Preço", self.styles['CommercialSubtitle']))
        
        table_data = [['Posição', 'Empresa', 'Preço Total', 'Diferença', 'Status', 'Score Comercial']]
        
        base_price = proposals_with_price[0]['preco_total']
        
        for i, proposal in enumerate(proposals_with_price, 1):
            if i == 1:
                diferenca = "Base"
                status = "🏆 Melhor Preço"
            else:
                diferenca_valor = proposal['preco_total'] - base_price
                diferenca_perc = ((proposal['preco_total'] / base_price) - 1) * 100
                diferenca = f"+R$ {diferenca_valor:,.2f}"
                status = f"📈 {diferenca_perc:.1f}% mais caro"
            
            table_data.append([
                f"<b>{i}º</b>",
                proposal['empresa'],
                f"<b>R$ {proposal['preco_total']:,.2f}</b>",
                diferenca,
                status,
                f"{proposal['score_comercial']:.1f}%"
            ])
        
        table = Table(table_data, colWidths=[1.5*cm, 3*cm, 2.5*cm, 2.5*cm, 3*cm, 2*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
        
        # Análise de dispersão de preços
        if len(proposals_with_price) > 1:
            story.append(Paragraph("📈 Análise de Dispersão de Preços", self.styles['CommercialSubtitle']))
            
            min_price = min(p['preco_total'] for p in proposals_with_price)
            max_price = max(p['preco_total'] for p in proposals_with_price)
            avg_price = sum(p['preco_total'] for p in proposals_with_price) / len(proposals_with_price)
            
            story.append(Paragraph(f"• <b>Menor Preço:</b> R$ {min_price:,.2f}", self.styles['CommercialNormal']))
            story.append(Paragraph(f"• <b>Maior Preço:</b> R$ {max_price:,.2f}", self.styles['CommercialNormal']))
            story.append(Paragraph(f"• <b>Preço Médio:</b> R$ {avg_price:,.2f}", self.styles['CommercialNormal']))
            story.append(Paragraph(f"• <b>Variação:</b> {((max_price/min_price - 1) * 100):.1f}%", self.styles['CommercialNormal']))
        
        story.append(self._create_separator_line())
    
    def _add_cost_analysis(self, story: List, proposals: List[Dict]):
        """Adiciona análise de custos e BDI"""
        section_title = Paragraph("💼 SEÇÃO 3: ANÁLISE DE CUSTOS E BDI", self.styles['CommercialSectionHeader'])
        story.append(section_title)
        
        proposals_with_price = [p for p in proposals if p['preco_total'] > 0]
        
        if not proposals_with_price:
            story.append(Paragraph("⚠️ Nenhuma proposta com informações comerciais para análise.", self.styles['CommercialNormal']))
            return
        
        for proposal in proposals_with_price:
            self._add_company_cost_analysis(story, proposal)
        
        story.append(self._create_separator_line())
    
    def _add_company_cost_analysis(self, story: List, proposal: Dict):
        """Adiciona análise de custos de uma empresa"""
        # Nome da empresa
        company_title = f"🏢 {proposal['empresa']} - Score Comercial: {proposal['score_comercial']:.1f}%"
        story.append(Paragraph(company_title, self.styles['CommercialSubtitle']))
        
        # Informações comerciais básicas
        story.append(Paragraph("💼 Informações Comerciais:", self.styles['CommercialNormal']))
        
        cnpj = proposal['cnpj'] if proposal['cnpj'] else "Não informado"
        story.append(Paragraph(f"• <b>CNPJ:</b> {cnpj}", self.styles['CommercialNormal']))
        story.append(Paragraph(f"• <b>Preço Total:</b> R$ {proposal['preco_total']:,.2f}", self.styles['CommercialNormal']))
        
        if proposal['bdi_percentual'] > 0:
            story.append(Paragraph(f"• <b>BDI:</b> {proposal['bdi_percentual']:.2f}%", self.styles['CommercialNormal']))
            
            # Análise do BDI
            if proposal['bdi_percentual'] <= 25:
                bdi_status = "✓ Excelente (≤ 25%)"
            elif proposal['bdi_percentual'] <= 35:
                bdi_status = "✓ Bom (≤ 35%)"
            elif proposal['bdi_percentual'] <= 45:
                bdi_status = "⚠️ Aceitável (≤ 45%)"
            else:
                bdi_status = "❌ Alto (> 45%)"
            story.append(Paragraph(f"• <b>Avaliação do BDI:</b> {bdi_status}", self.styles['CommercialNormal']))
        else:
            story.append(Paragraph("• <b>BDI:</b> Não informado", self.styles['CommercialNormal']))
        
        # Composição de custos
        if any(proposal['composicao_custos'].values()):
            story.append(Paragraph("💰 Composição de Custos:", self.styles['CommercialNormal']))
            
            total_custos = sum(proposal['composicao_custos'].values())
            
            for categoria, valor in proposal['composicao_custos'].items():
                if valor > 0:
                    percentual = (valor / total_custos) * 100 if total_custos > 0 else 0
                    categoria_nome = categoria.replace('_', ' ').title()
                    story.append(Paragraph(f"• <b>{categoria_nome}:</b> R$ {valor:,.2f} ({percentual:.1f}%)", self.styles['CommercialNormal']))
            
            story.append(Paragraph(f"• <b>Total de Custos Diretos:</b> R$ {total_custos:,.2f}", self.styles['CommercialNormal']))
            
            # Calcular BDI implícito se não informado
            if proposal['bdi_percentual'] == 0 and total_custos > 0:
                bdi_implicito = ((proposal['preco_total'] / total_custos) - 1) * 100
                story.append(Paragraph(f"• <b>BDI Implícito:</b> {bdi_implicito:.2f}%", self.styles['CommercialNormal']))
        
        # Condições comerciais
        if proposal['condicoes_pagamento']:
            story.append(Paragraph(f"• <b>Condições de Pagamento:</b> {proposal['condicoes_pagamento']}", self.styles['CommercialNormal']))
        
        if proposal['garantia']:
            story.append(Paragraph(f"• <b>Garantia:</b> {proposal['garantia']}", self.styles['CommercialNormal']))
        
        story.append(Spacer(1, 15))
    
    def _add_commercial_conditions(self, story: List, proposals: List[Dict]):
        """Adiciona análise de condições comerciais"""
        section_title = Paragraph("📋 SEÇÃO 4: CONDIÇÕES COMERCIAIS", self.styles['CommercialSectionHeader'])
        story.append(section_title)
        
        # Tabela comparativa de condições
        story.append(Paragraph("📊 Comparativo de Condições Comerciais", self.styles['CommercialSubtitle']))
        
        proposals_with_price = [p for p in proposals if p['preco_total'] > 0]
        
        if proposals_with_price:
            table_data = [['Empresa', 'Preço Total', 'BDI (%)', 'Pagamento', 'Garantia']]
            
            for proposal in proposals_with_price:
                pagamento = proposal['condicoes_pagamento'] if proposal['condicoes_pagamento'] else "Não informado"
                garantia = proposal['garantia'] if proposal['garantia'] else "Não informado"
                bdi = f"{proposal['bdi_percentual']:.1f}%" if proposal['bdi_percentual'] > 0 else "N/I"
                
                table_data.append([
                    proposal['empresa'],
                    f"R$ {proposal['preco_total']:,.2f}",
                    bdi,
                    pagamento[:30] + "..." if len(pagamento) > 30 else pagamento,
                    garantia[:20] + "..." if len(garantia) > 20 else garantia
                ])
            
            table = Table(table_data, colWidths=[3*cm, 2.5*cm, 1.5*cm, 4*cm, 3*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkred),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.mistyrose),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            
            story.append(table)
        
        story.append(self._create_separator_line())
    
    def _add_commercial_recommendations(self, story: List, analysis_result: Dict):
        """Adiciona recomendações comerciais"""
        section_title = Paragraph("💡 SEÇÃO 5: RECOMENDAÇÕES COMERCIAIS", self.styles['CommercialSectionHeader'])
        story.append(section_title)
        
        proposals = analysis_result['proposals']
        proposals_with_price = [p for p in proposals if p['preco_total'] > 0]
        
        if proposals_with_price:
            # Ordenar por score comercial
            proposals_with_price.sort(key=lambda x: x['score_comercial'], reverse=True)
            best_commercial = proposals_with_price[0]
            
            # Recomendação principal
            story.append(Paragraph("🏆 Recomendação Comercial Principal", self.styles['CommercialSubtitle']))
            story.append(Paragraph(f"Com base na análise comercial, recomenda-se a empresa <b>{best_commercial['empresa']}</b> que obteve o melhor score comercial ({best_commercial['score_comercial']:.1f}%) com preço de R$ {best_commercial['preco_total']:,.2f}.", self.styles['CommercialNormal']))
            
            # Justificativa comercial
            story.append(Paragraph("📋 Justificativa Comercial:", self.styles['CommercialSubtitle']))
            
            # Encontrar posição no ranking de preços
            proposals_by_price = sorted(proposals_with_price, key=lambda x: x['preco_total'])
            price_position = proposals_by_price.index(best_commercial) + 1
            
            story.append(Paragraph(f"• <b>Posição no ranking de preços:</b> {price_position}º lugar", self.styles['CommercialNormal']))
            story.append(Paragraph(f"• <b>Preço proposto:</b> R$ {best_commercial['preco_total']:,.2f}", self.styles['CommercialNormal']))
            
            if best_commercial['bdi_percentual'] > 0:
                story.append(Paragraph(f"• <b>BDI oferecido:</b> {best_commercial['bdi_percentual']:.2f}%", self.styles['CommercialNormal']))
            
            # Análise custo-benefício
            if price_position == 1:
                story.append(Paragraph("• <b>Vantagem:</b> Melhor preço do certame", self.styles['CommercialNormal']))
            else:
                cheapest = proposals_by_price[0]
                difference = best_commercial['preco_total'] - cheapest['preco_total']
                percentage = (difference / cheapest['preco_total']) * 100
                story.append(Paragraph(f"• <b>Diferença para o menor preço:</b> R$ {difference:,.2f} ({percentage:.1f}%)", self.styles['CommercialNormal']))
        
        # Ações comerciais recomendadas
        story.append(Paragraph("💼 Ações Comerciais Recomendadas", self.styles['CommercialSubtitle']))
        actions = [
            "• Negociar melhores condições de pagamento com as empresas finalistas",
            "• Solicitar detalhamento da composição de custos",
            "• Verificar referências comerciais das empresas",
            "• Confirmar capacidade financeira para execução do projeto",
            "• Negociar extensão do período de garantia",
            "• Definir critérios de reajuste de preços",
            "• Estabelecer multas por atraso na entrega",
            "• Avaliar propostas de desconto para pagamento antecipado"
        ]
        
        for action in actions:
            story.append(Paragraph(action, self.styles['CommercialNormal']))
        
        # Próximos passos comerciais
        story.append(Paragraph("🚀 Próximos Passos Comerciais", self.styles['CommercialSubtitle']))
        next_steps = [
            "• Reunião comercial com as empresas finalistas",
            "• Negociação de condições específicas",
            "• Verificação de documentação fiscal",
            "• Análise de capacidade de execução financeira",
            "• Definição de cronograma de pagamentos",
            "• Elaboração de minuta contratual"
        ]
        
        for step in next_steps:
            story.append(Paragraph(step, self.styles['CommercialNormal']))
    
    def _create_separator_line(self):
        """Cria linha separadora"""
        return Table([['---']], colWidths=[15*cm], style=TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 14),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.grey)
        ]))

# Instanciar analisador e geradores
analyzer = ProposalAnalyzer()
technical_report_generator = TechnicalReportGenerator()
commercial_report_generator = CommercialReportGenerator()

def allowed_file(filename):
    """Verifica se o arquivo tem extensão permitida"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    """Página principal"""
    return render_template_string(HTML_TEMPLATE)

@app.route('/upload', methods=['POST'])
def upload_files():
    """Endpoint para upload e processamento de arquivos"""
    try:
        if 'files' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        files = request.files.getlist('files')
        report_type = request.form.get('report_type', 'both')  # 'technical', 'commercial', 'both'
        
        if not files or all(file.filename == '' for file in files):
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        uploaded_files = []
        
        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"{timestamp}_{filename}"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                
                file.save(filepath)
                
                file_extension = filename.rsplit('.', 1)[1].lower()
                uploaded_files.append({
                    'filename': filename,
                    'path': filepath,
                    'type': 'pdf' if file_extension == 'pdf' else file_extension
                })
        
        if not uploaded_files:
            return jsonify({'error': 'Nenhum arquivo válido foi enviado'}), 400
        
        # Processar arquivos
        logger.info(f"Processando {len(uploaded_files)} arquivos")
        analysis_result = analyzer.analyze_proposals(uploaded_files)
        
        # Gerar relatórios conforme solicitado
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        report_urls = []
        
        if report_type in ['technical', 'both']:
            technical_filename = f'analise_tecnica_{timestamp}.pdf'
            technical_path = os.path.join(app.config['UPLOAD_FOLDER'], technical_filename)
            technical_report_generator.generate_technical_report(analysis_result, technical_path)
            report_urls.append({
                'type': 'technical',
                'filename': technical_filename,
                'url': f'/download/{technical_filename}',
                'title': 'Relatório de Análise Técnica'
            })
        
        if report_type in ['commercial', 'both']:
            commercial_filename = f'analise_comercial_{timestamp}.pdf'
            commercial_path = os.path.join(app.config['UPLOAD_FOLDER'], commercial_filename)
            commercial_report_generator.generate_commercial_report(analysis_result, commercial_path)
            report_urls.append({
                'type': 'commercial',
                'filename': commercial_filename,
                'url': f'/download/{commercial_filename}',
                'title': 'Relatório de Análise Comercial'
            })
        
        # Limpar arquivos temporários
        for file_info in uploaded_files:
            try:
                os.remove(file_info['path'])
            except:
                pass
        
        return jsonify({
            'success': True,
            'message': 'Análise concluída com sucesso!',
            'reports': report_urls,
            'summary': analysis_result['summary']
        })
        
    except Exception as e:
        logger.error(f"Erro no processamento: {str(e)}")
        return jsonify({'error': f'Erro no processamento: {str(e)}'}), 500

@app.route('/download/<filename>')
def download_file(filename):
    """Endpoint para download do relatório"""
    try:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)
        else:
            return jsonify({'error': 'Arquivo não encontrado'}), 404
    except Exception as e:
        logger.error(f"Erro no download: {str(e)}")
        return jsonify({'error': 'Erro no download'}), 500



# Template HTML atualizado para escolher tipo de relatório
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Arias Analyzer Pro - Análise de Propostas</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 20px;
        }
        
        .container {
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 40px rgba(0,0,0,0.1);
            padding: 40px;
            max-width: 600px;
            width: 100%;
            text-align: center;
        }
        
        .logo {
            font-size: 2.5em;
            font-weight: bold;
            color: #667eea;
            margin-bottom: 10px;
        }
        
        .subtitle {
            color: #666;
            margin-bottom: 30px;
            font-size: 1.1em;
        }
        
        .upload-area {
            border: 3px dashed #667eea;
            border-radius: 15px;
            padding: 40px 20px;
            margin: 30px 0;
            background: #f8f9ff;
            transition: all 0.3s ease;
            cursor: pointer;
        }
        
        .upload-area:hover {
            border-color: #764ba2;
            background: #f0f2ff;
        }
        
        .upload-area.dragover {
            border-color: #764ba2;
            background: #e8ebff;
            transform: scale(1.02);
        }
        
        .upload-icon {
            font-size: 3em;
            color: #667eea;
            margin-bottom: 15px;
        }
        
        .upload-text {
            color: #333;
            font-size: 1.1em;
            margin-bottom: 10px;
        }
        
        .upload-hint {
            color: #666;
            font-size: 0.9em;
        }
        
        .file-input {
            display: none;
        }
        
        .report-type-section {
            margin: 20px 0;
            padding: 20px;
            background: #f8f9ff;
            border-radius: 10px;
            border: 1px solid #e0e6ff;
        }
        
        .report-type-title {
            font-size: 1.2em;
            color: #333;
            margin-bottom: 15px;
            font-weight: 600;
        }
        
        .report-options {
            display: flex;
            gap: 15px;
            justify-content: center;
            flex-wrap: wrap;
        }
        
        .report-option {
            display: flex;
            align-items: center;
            gap: 8px;
            padding: 10px 15px;
            background: white;
            border: 2px solid #e0e6ff;
            border-radius: 8px;
            cursor: pointer;
            transition: all 0.3s ease;
        }
        
        .report-option:hover {
            border-color: #667eea;
            background: #f0f2ff;
        }
        
        .report-option input[type="radio"] {
            margin: 0;
        }
        
        .report-option.selected {
            border-color: #667eea;
            background: #667eea;
            color: white;
        }
        
        .analyze-btn {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            padding: 15px 40px;
            border-radius: 50px;
            font-size: 1.1em;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s ease;
            margin-top: 20px;
            width: 100%;
            max-width: 300px;
        }
        
        .analyze-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 10px 20px rgba(102, 126, 234, 0.3);
        }
        
        .analyze-btn:disabled {
            background: #ccc;
            cursor: not-allowed;
            transform: none;
            box-shadow: none;
        }
        
        .progress {
            display: none;
            margin: 20px 0;
        }
        
        .progress-bar {
            width: 100%;
            height: 8px;
            background: #e0e6ff;
            border-radius: 4px;
            overflow: hidden;
        }
        
        .progress-fill {
            height: 100%;
            background: linear-gradient(90deg, #667eea, #764ba2);
            width: 0%;
            transition: width 0.3s ease;
            animation: pulse 2s infinite;
        }
        
        @keyframes pulse {
            0%, 100% { opacity: 1; }
            50% { opacity: 0.7; }
        }
        
        .file-list {
            margin: 20px 0;
            text-align: left;
        }
        
        .file-item {
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding: 10px;
            background: #f8f9ff;
            border-radius: 8px;
            margin: 5px 0;
            border: 1px solid #e0e6ff;
        }
        
        .file-name {
            color: #333;
            font-weight: 500;
        }
        
        .file-size {
            color: #666;
            font-size: 0.9em;
        }
        
        .remove-file {
            background: #ff4757;
            color: white;
            border: none;
            border-radius: 50%;
            width: 25px;
            height: 25px;
            cursor: pointer;
            font-size: 0.8em;
        }
        
        .results {
            display: none;
            margin-top: 30px;
            padding: 20px;
            background: #f8f9ff;
            border-radius: 15px;
            border: 1px solid #e0e6ff;
        }
        
        .success-icon {
            font-size: 3em;
            color: #2ed573;
            margin-bottom: 15px;
        }
        
        .download-links {
            margin: 20px 0;
        }
        
        .download-btn {
            display: inline-block;
            background: #2ed573;
            color: white;
            text-decoration: none;
            padding: 12px 25px;
            border-radius: 25px;
            margin: 5px;
            font-weight: 600;
            transition: all 0.3s ease;
        }
        
        .download-btn:hover {
            background: #26d467;
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(46, 213, 115, 0.3);
        }
        
        .download-btn.technical {
            background: #3742fa;
        }
        
        .download-btn.technical:hover {
            background: #2f3542;
        }
        
        .download-btn.commercial {
            background: #ff6b6b;
        }
        
        .download-btn.commercial:hover {
            background: #ff5252;
        }
        
        .error {
            color: #ff4757;
            background: #ffe0e0;
            padding: 15px;
            border-radius: 8px;
            margin: 15px 0;
            border: 1px solid #ffcdd2;
        }
        
        .azure-status {
            margin: 15px 0;
            padding: 10px;
            border-radius: 8px;
            font-size: 0.9em;
        }
        
        .azure-active {
            background: #e8f5e8;
            color: #2e7d32;
            border: 1px solid #c8e6c9;
        }
        
        .azure-inactive {
            background: #fff3e0;
            color: #f57c00;
            border: 1px solid #ffcc02;
        }
        
        @media (max-width: 768px) {
            .container {
                padding: 20px;
                margin: 10px;
            }
            
            .logo {
                font-size: 2em;
            }
            
            .report-options {
                flex-direction: column;
                align-items: center;
            }
            
            .report-option {
                width: 100%;
                max-width: 250px;
                justify-content: center;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="logo">📊 Arias Analyzer Pro</div>
        <div class="subtitle">Análise Inteligente de Propostas com Azure AI</div>
        
        <div class="azure-status" id="azureStatus">
            🤖 Azure Document Intelligence: Ativo
        </div>
        
        <form id="uploadForm" enctype="multipart/form-data">
            <div class="upload-area" id="uploadArea">
                <div class="upload-icon">📁</div>
                <div class="upload-text">Clique aqui ou arraste os arquivos</div>
                <div class="upload-hint">PDFs e planilhas Excel (máx. 50MB cada)</div>
                <input type="file" id="fileInput" name="files" multiple accept=".pdf,.xlsx,.xls" class="file-input">
            </div>
            
            <div class="file-list" id="fileList"></div>
            
            <div class="report-type-section">
                <div class="report-type-title">🎯 Tipo de Relatório</div>
                <div class="report-options">
                    <label class="report-option" for="reportBoth">
                        <input type="radio" id="reportBoth" name="report_type" value="both" checked>
                        <span>📋 Ambos os Relatórios</span>
                    </label>
                    <label class="report-option" for="reportTechnical">
                        <input type="radio" id="reportTechnical" name="report_type" value="technical">
                        <span>🔧 Apenas Técnico</span>
                    </label>
                    <label class="report-option" for="reportCommercial">
                        <input type="radio" id="reportCommercial" name="report_type" value="commercial">
                        <span>💰 Apenas Comercial</span>
                    </label>
                </div>
            </div>
            
            <button type="submit" class="analyze-btn" id="analyzeBtn" disabled>
                🚀 Analisar Propostas
            </button>
        </form>
        
        <div class="progress" id="progress">
            <div class="progress-bar">
                <div class="progress-fill" id="progressFill"></div>
            </div>
            <div style="margin-top: 10px; color: #667eea; font-weight: 600;" id="progressText">
                Processando arquivos...
            </div>
        </div>
        
        <div class="results" id="results"></div>
    </div>

    <script>
        const uploadArea = document.getElementById('uploadArea');
        const fileInput = document.getElementById('fileInput');
        const fileList = document.getElementById('fileList');
        const analyzeBtn = document.getElementById('analyzeBtn');
        const uploadForm = document.getElementById('uploadForm');
        const progress = document.getElementById('progress');
        const progressFill = document.getElementById('progressFill');
        const progressText = document.getElementById('progressText');
        const results = document.getElementById('results');
        
        let selectedFiles = [];
        
        // Eventos de drag and drop
        uploadArea.addEventListener('click', () => fileInput.click());
        uploadArea.addEventListener('dragover', (e) => {
            e.preventDefault();
            uploadArea.classList.add('dragover');
        });
        uploadArea.addEventListener('dragleave', () => {
            uploadArea.classList.remove('dragover');
        });
        uploadArea.addEventListener('drop', (e) => {
            e.preventDefault();
            uploadArea.classList.remove('dragover');
            handleFiles(e.dataTransfer.files);
        });
        
        fileInput.addEventListener('change', (e) => {
            handleFiles(e.target.files);
        });
        
        // Gerenciar seleção de tipo de relatório
        document.querySelectorAll('input[name="report_type"]').forEach(radio => {
            radio.addEventListener('change', updateReportSelection);
        });
        
        function updateReportSelection() {
            document.querySelectorAll('.report-option').forEach(option => {
                option.classList.remove('selected');
            });
            
            const selected = document.querySelector('input[name="report_type"]:checked');
            if (selected) {
                selected.closest('.report-option').classList.add('selected');
            }
        }
        
        // Inicializar seleção
        updateReportSelection();
        
        function handleFiles(files) {
            Array.from(files).forEach(file => {
                if (file.size > 50 * 1024 * 1024) {
                    alert(`Arquivo ${file.name} é muito grande (máx. 50MB)`);
                    return;
                }
                
                if (!file.name.match(/\.(pdf|xlsx|xls)$/i)) {
                    alert(`Arquivo ${file.name} não é suportado`);
                    return;
                }
                
                if (!selectedFiles.find(f => f.name === file.name)) {
                    selectedFiles.push(file);
                }
            });
            
            updateFileList();
            updateAnalyzeButton();
        }
        
        function updateFileList() {
            fileList.innerHTML = '';
            
            selectedFiles.forEach((file, index) => {
                const fileItem = document.createElement('div');
                fileItem.className = 'file-item';
                fileItem.innerHTML = `
                    <div>
                        <div class="file-name">${file.name}</div>
                        <div class="file-size">${formatFileSize(file.size)}</div>
                    </div>
                    <button type="button" class="remove-file" onclick="removeFile(${index})">×</button>
                `;
                fileList.appendChild(fileItem);
            });
        }
        
        function removeFile(index) {
            selectedFiles.splice(index, 1);
            updateFileList();
            updateAnalyzeButton();
        }
        
        function updateAnalyzeButton() {
            analyzeBtn.disabled = selectedFiles.length === 0;
        }
        
        function formatFileSize(bytes) {
            if (bytes === 0) return '0 Bytes';
            const k = 1024;
            const sizes = ['Bytes', 'KB', 'MB', 'GB'];
            const i = Math.floor(Math.log(bytes) / Math.log(k));
            return parseFloat((bytes / Math.pow(k, i)).toFixed(2)) + ' ' + sizes[i];
        }
        
        uploadForm.addEventListener('submit', async (e) => {
            e.preventDefault();
            
            if (selectedFiles.length === 0) {
                alert('Selecione pelo menos um arquivo');
                return;
            }
            
            const formData = new FormData();
            selectedFiles.forEach(file => {
                formData.append('files', file);
            });
            
            const reportType = document.querySelector('input[name="report_type"]:checked').value;
            formData.append('report_type', reportType);
            
            // Mostrar progresso
            progress.style.display = 'block';
            results.style.display = 'none';
            analyzeBtn.disabled = true;
            
            let progressValue = 0;
            const progressInterval = setInterval(() => {
                progressValue += Math.random() * 15;
                if (progressValue > 90) progressValue = 90;
                progressFill.style.width = progressValue + '%';
            }, 500);
            
            try {
                const response = await fetch('/upload', {
                    method: 'POST',
                    body: formData
                });
                
                const result = await response.json();
                
                clearInterval(progressInterval);
                progressFill.style.width = '100%';
                
                setTimeout(() => {
                    progress.style.display = 'none';
                    
                    if (result.success) {
                        showResults(result);
                    } else {
                        showError(result.error || 'Erro desconhecido');
                    }
                    
                    analyzeBtn.disabled = false;
                }, 1000);
                
            } catch (error) {
                clearInterval(progressInterval);
                progress.style.display = 'none';
                showError('Erro de conexão: ' + error.message);
                analyzeBtn.disabled = false;
            }
        });
        
        function showResults(result) {
            let downloadLinks = '';
            
            result.reports.forEach(report => {
                const btnClass = report.type === 'technical' ? 'technical' : 
                               report.type === 'commercial' ? 'commercial' : '';
                downloadLinks += `
                    <a href="${report.url}" class="download-btn ${btnClass}" target="_blank">
                        📄 ${report.title}
                    </a>
                `;
            });
            
            results.innerHTML = `
                <div class="success-icon">✅</div>
                <h3 style="color: #2ed573; margin-bottom: 15px;">Análise Concluída!</h3>
                <p style="margin-bottom: 20px;">
                    ${result.reports.length} relatório(s) gerado(s) com sucesso.
                </p>
                <div class="download-links">
                    ${downloadLinks}
                </div>
                <div style="margin-top: 20px; font-size: 0.9em; color: #666;">
                    <strong>Resumo:</strong><br>
                    • Total de propostas: ${result.summary.total_proposals || 0}<br>
                    ${result.summary.best_technical ? `• Melhor técnica: ${result.summary.best_technical}<br>` : ''}
                    ${result.summary.best_commercial ? `• Melhor comercial: ${result.summary.best_commercial}` : ''}
                </div>
            `;
            results.style.display = 'block';
        }
        
        function showError(message) {
            results.innerHTML = `
                <div class="error">
                    <strong>❌ Erro no processamento:</strong><br>
                    ${message}
                </div>
            `;
            results.style.display = 'block';
        }
    </script>
</body>
</html>
'''

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)

